import re
import time
import random
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://www.gartner.com/reviews/product/bmc-helix-itsm"
MAX_PAGES = 30
OUTPUT_FILE = "gartner_bmc_helix_reviews.xlsx"

COOKIE = """BISKIT"""
HEADERS = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-encoding": "gzip, deflate, br",
    "accept-language": "en-US,en;q=0.9,hi;q=0.8",
    "cache-control": "max-age=0",
    "cookie": COOKIE,
    "sec-ch-ua": '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
    "sec-ch-ua-arch": '""',
    "sec-ch-ua-bitness": '"64"',
    "sec-ch-ua-full-version": '"110.0.5481.178"',
    "sec-ch-ua-full-version-list": '"Chromium";v="110.0.5481.178", "Not A(Brand";v="24.0.0.0", "Google Chrome";v="110.0.5481.178"',
    "sec-ch-ua-mobile": "?1",
    "sec-ch-ua-model": '"Nexus 5"',
    "sec-ch-ua-platform": '"Android"',
    "sec-ch-ua-platform-version": '"6.0"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "same-origin",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Mobile Safari/537.36",
}

def get_rating(el):
    if not el:
        return ""
    for attr in ["aria-label", "title", "data-score", "data-rating"]:
        val = el.get(attr, "")
        m = re.search(r"(\d+(\.\d+)?)\s*(?:out of|/)\s*5", val, re.IGNORECASE)
        if m:
            return float(m.group(1))
    text = el.get_text()
    m = re.search(r"\b([1-5](\.\d)?)\b", text)
    return float(m.group(1)) if m else ""


def parse_page(soup, page_num):
    reviews = []

    if page_num == 1:
        with open("debug_page1.html", "w", encoding="utf-8") as f:
            f.write(str(soup))
        print("Saved debug_page1.html")

    containers = []
    for attempt in [
        lambda: soup.find_all("div", class_=re.compile(r"review-content|reviewContent", re.I)),
        lambda: soup.find_all("div", class_=re.compile(r"reviewCard|review-card|ReviewCard|review_card", re.I)),
        lambda: soup.find_all("article", class_=re.compile(r"review", re.I)),
        lambda: soup.find_all(attrs={"data-testid": re.compile(r"review", re.I)}),
        lambda: soup.find_all("div", class_=re.compile(r"userReview|peer-review|rt-review", re.I)),
        lambda: soup.find_all("li", class_=re.compile(r"review", re.I)),
        lambda: [d for d in soup.find_all("div", recursive=False)
                 if d.find(attrs={"aria-label": re.compile(r"out of 5", re.I)})
                 and len(d.get_text()) > 150],
    ]:
        containers = attempt()
        if containers:
            print(f"  📦 {len(containers)} review containers found")
            break

    if not containers:
        snippet = soup.get_text()[:400].replace("\n", " ")
        print(f"No containers matched. Page text: {snippet}")
        return []

    for card in containers:
        r = {
            "title": "", "date": "", "overall_rating": "",
            "product": "BMC Helix ITSM", "overall_comment": "",
            "integration_and_deployment": "", "service_and_support": "",
            "product_capabilities": "", "role": "", "industry": "",
            "function": "", "firm_size": "", "deployment": "",
        }
        text = card.get_text("\n", strip=True)
        for tag in ["h3", "h2", "h4", "h5"]:
            el = card.find(tag)
            if el and len(el.get_text(strip=True)) > 5:
                r["title"] = el.get_text(strip=True).strip('"').strip()
                break
        if not r["title"]:
            el = card.find(class_=re.compile(r"title|headline|subject", re.I))
            if el:
                r["title"] = el.get_text(strip=True).strip('"').strip()

        # Date
        el = card.find("time")
        if el:
            r["date"] = el.get("datetime", el.get_text(strip=True))
        else:
            el = card.find(class_=re.compile(r"date|submitted|published|when", re.I))
            if el:
                r["date"] = el.get_text(strip=True)
            else:
                m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+\d{1,2},?\s+\d{4}", text)
                if m:
                    r["date"] = m.group()

        for el in card.find_all(attrs={"aria-label": re.compile(r"out of 5", re.I)}):
            val = get_rating(el)
            if val:
                r["overall_rating"] = val
                break
        if not r["overall_rating"]:
            el = card.find(class_=re.compile(r"overall.?rating|rating.?overall|starRating|star-rating", re.I))
            r["overall_rating"] = get_rating(el)

        for cls in [r"comment|review.?body|reviewBody|overall.?comment", r"description|summary|excerpt"]:
            el = card.find(class_=re.compile(cls, re.I))
            if el and len(el.get_text(strip=True)) > 40:
                r["overall_comment"] = el.get_text(strip=True)
                break
        if not r["overall_comment"]:
            paras = sorted(
                [p.get_text(strip=True) for p in card.find_all("p") if len(p.get_text(strip=True)) > 60],
                key=len, reverse=True
            )
            if paras:
                r["overall_comment"] = paras[0]

        for key, kws in {
            "integration_and_deployment": ["Integration & Deployment", "Integration and Deployment", "Integration"],
            "service_and_support": ["Service & Support", "Service and Support", "Support"],
            "product_capabilities": ["Product Capabilities", "Capabilities"],
        }.items():
            for kw in kws:
                label_el = card.find(string=re.compile(re.escape(kw), re.I))
                if label_el:
                    parent = label_el.find_parent()
                    if parent:
                        rating_el = parent.find_next(attrs={"aria-label": re.compile(r"out of 5", re.I)})
                        if not rating_el:
                            rating_el = parent.find_next(class_=re.compile(r"star|rating|score", re.I))
                        val = get_rating(rating_el)
                        if val:
                            r[key] = val
                            break
                m = re.search(rf"{re.escape(kw)}[^\d]{{0,40}}(\d(\.\d)?)", text, re.IGNORECASE)
                if m:
                    r[key] = float(m.group(1))
                    break

        profile_el = card.find(class_=re.compile(r"profile|reviewer|author|contributor|sidebar", re.I))
        profile_text = profile_el.get_text("\n") if profile_el else text

        field_patterns = {
            "role": [
                r"(?:Role|Job Title|Title|Position)[:\s]+([^\n|,<]{3,60})",
                r"\b(IT Manager|Director|Analyst|Engineer|Architect|CTO|CIO|CISO|VP|AVP|Manager|Consultant|Administrator|Specialist|Lead|Head of IT|IT Director)[^\n]{0,30}",
            ],
            "industry": [
                r"(?:Industry)[:\s]+([^\n|,<]{3,50})",
                r"\b(Retail|Healthcare|Finance|Banking|Insurance|Manufacturing|Technology|Government|Education|Telecommunications|Energy|Transportation|Media|Pharmaceuticals|Automotive|Construction)\b",
            ],
            "function": [
                r"(?:Function)[:\s]+([^\n|,<]{2,40})",
                r"\b(IT|Finance|Operations|HR|Marketing|Sales|Legal|Engineering|Procurement|Security)\b",
            ],
            "firm_size": [
                r"(\$?\d+[BMK]?\s*[-–]\s*\$?\d+\s*(?:Billion|Million|[BMK])?\s*(?:USD)?)",
                r"(\d+\s*[-–]\s*\d+\s*(?:employees|staff|people))",
            ],
            "deployment": [
                r"(?:Deployment(?:\s+Architecture)?)[:\s]+([^\n|,<]{3,70})",
                r"\b(Cloud|SaaS|PaaS|IaaS|On-Premise|On-Prem|Hybrid|Private Cloud|Public Cloud|Self-Hosted)\b",
            ],
        }

        for key, pats in field_patterns.items():
            for pat in pats:
                m = re.search(pat, profile_text, re.IGNORECASE)
                if m:
                    r[key] = m.group(1).strip()
                    break

        if r["title"] or r["overall_comment"]:
            reviews.append(r)

    return reviews


def scrape():
    all_reviews = []

    try:
        from curl_cffi import requests as cffi_requests
        session = cffi_requests.Session(impersonate="chrome110")
        use_cffi = True
        print("Using curl_cffi (Chrome TLS fingerprint)")
    except ImportError:
        import requests
        session = requests.Session()
        use_cffi = False
        print("curl_cffi not found — using requests")
        print("   Run: pip install curl_cffi  for better results\n")

    for page_num in range(1, MAX_PAGES + 1):
        url = f"{BASE_URL}?page={page_num}"
        print(f"\n📄 Page {page_num}: {url}")

        referer = BASE_URL if page_num > 1 else "https://www.gartner.com/reviews/market/it-service-management-tools"
        h = {**HEADERS, "referer": referer}

        try:
            resp = session.get(url, headers=h, timeout=30)
            print(f"  Status: {resp.status_code} | Size: {len(resp.text):,} chars")

            if resp.status_code == 403:
                print("   403 — Cookies expired. Re-copy cookies from Chrome F12 > Network.")
                print("     Update the COOKIE variable at the top of this script.")
                break
            if resp.status_code == 429:
                print("   Rate limited. Waiting 60s...")
                time.sleep(60)
                continue
            if resp.status_code != 200:
                print(f"   Unexpected status {resp.status_code}")
                break
            if "challenge-platform" in resp.text or "Just a moment" in resp.text:
                print("   Cloudflare block. Cookies expired — re-copy from Chrome.")
                break

            soup = BeautifulSoup(resp.text, "html.parser")
            page_reviews = parse_page(soup, page_num)

            if not page_reviews:
                print(f"   No reviews on page {page_num} — done.")
                break

            all_reviews.extend(page_reviews)
            print(f"   {len(page_reviews)} reviews | Total: {len(all_reviews)}")

            # Check for next page button
            next_btn = soup.find(attrs={"aria-label": re.compile(r"next page|next", re.I)})
            if next_btn and next_btn.get("disabled"):
                print("  Last page reached.")
                break

            delay = random.uniform(4, 8)
            print(f"   Waiting {delay:.1f}s...")
            time.sleep(delay)

        except Exception as e:
            print(f"  ❌ Error: {e}")
            break

    return all_reviews

def save_excel(reviews):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gartner Reviews"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    alt_fill = PatternFill("solid", start_color="DCE6F1")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Review Title", "Date Submitted", "Overall Rating (1-5)", "Product",
        "Overall Comment", "Integration & Deployment", "Service & Support",
        "Product Capabilities", "Reviewer Role", "Industry", "Function",
        "Firm Size", "Deployment Architecture",
    ]
    widths = [42, 16, 20, 20, 70, 23, 18, 23, 23, 20, 15, 22, 30]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    keys = [
        "title", "date", "overall_rating", "product", "overall_comment",
        "integration_and_deployment", "service_and_support", "product_capabilities",
        "role", "industry", "function", "firm_size", "deployment",
    ]
    for ri, rev in enumerate(reviews, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, key in enumerate(keys, 1):
            c = ws.cell(row=ri, column=ci, value=rev.get(key, ""))
            c.alignment = wrap
            c.border = thin
            if fill:
                c.fill = fill
        ws.row_dimensions[ri].height = 70

    ws2 = wb.create_sheet("Summary")
    for r, (k, v) in enumerate([
        ("Total Reviews Scraped", len(reviews)),
        ("Source URL", BASE_URL),
        ("Product", "BMC Helix ITSM"),
        ("Scrape Method", "Browser Cookie + curl_cffi"),
    ], 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
        ws2.cell(row=r, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 65

    wb.save(OUTPUT_FILE)
    print(f"\n Saved {len(reviews)} reviews → {OUTPUT_FILE}")


if __name__ == "__main__":
    import subprocess, sys
    try:
        import curl_cffi
    except ImportError:
        print("Installing curl_cffi...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "curl_cffi", "-q"])

    print("=" * 60)
    print("  Gartner BMC Helix ITSM Scraper")
    print("=" * 60)

    reviews = scrape()

    if reviews:
        save_excel(reviews)
        print(f"\n🎉 Done! Open: {OUTPUT_FILE}")
    else:
        print("\n No reviews scraped.")
        print("   → If cookies expired, re-copy from Chrome F12 > Network > cookie header")